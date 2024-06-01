import openpyxl
# Give the location of the file
my_path = "Book1.xlsx"
# workbook object is created
my_wb_obj = openpyxl.load_workbook(my_path)
my_sheet_obj = my_wb_obj.active
my_max_col = my_sheet_obj.max_column
for i in range(1, my_max_col + 1):
   my_cell_obj = my_sheet_obj.cell(row = 2, column = i)
   print(my_cell_obj.value)
