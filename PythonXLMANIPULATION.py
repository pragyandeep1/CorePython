import openpyxl

my_wb = openpyxl.Workbook()
my_sheet = my_wb.active
my_wb.create_sheet(index = 2 , title = "Puneeth")
my_wb.save("Book3.xlsx")

my_path = "Book2.xlsx"
my_wb_obj = openpyxl.load_workbook(my_path)
my_sheet_obj = my_wb_obj.active
print("maxm row "+my_sheet_obj.max_row.__str__())
my_cell_obj = my_sheet_obj.cell(row = 1, column = 2)
print(my_cell_obj.value)
my_sheet_obj = my_wb_obj.active
print("Max col "+my_sheet_obj.max_column.__str__())
